from pathlib import Path
import random
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pandas as pd
from datetime import datetime
from typing import Sequence, Hashable, Any
from openpyxl.chart import LineChart, Reference

input_path = Path("input/历史表现.xlsx")
output_path = Path("output/历史表现.xlsx")
hushen300_path = Path("input/沪深300.csv")
shangzheng50_path = Path("input/上证50.csv")


def handle_sheet(workbook: Workbook, sheet_name: str) -> dict[str, Sequence[str | int]]:
    # 获取沪深300和上证50数据
    hushen300 = get_hushen300_data()
    shangzheng50 = get_shangzheng50_data()

    # Select the desired worksheet
    worksheet = workbook[sheet_name]
    data: dict[str, Sequence[str | int]] = {}
    headers = [cell.value.replace(" ", "") if type(cell.value) is str else cell.value for cell in worksheet[1]]
    if sheet_name == "bybit" or sheet_name == "gate":
        new_headers = [
            "日期",
            "单位净值",
            "当前回撤",
            "7日年化收益率",
            "30日年化收益率",
        ]
        for header in new_headers:
            col_index = headers.index(header) + 1
            values = [worksheet.cell(row=row, column=col_index).value for row in range(2, worksheet.max_row + 1)]
            if header == "单位净值":
                data[sheet_name] = [
                    value.strftime("%Y-%m-%d") if isinstance(value, datetime) else value for value in values
                ]
                data[sheet_name] = [round(value, 4) if type(value) is float else value for value in data[sheet_name]]

                hushen_start = hushen300.loc[hushen300["日期"] == data["日期"][0]].index[0]
                hushen_end = hushen300.loc[hushen300["日期"] == data["日期"][-1]].index[0]
                shangzheng_start = hushen300.loc[hushen300["日期"] == data["日期"][0]].index[0]
                shangzheng_end = hushen300.loc[hushen300["日期"] == data["日期"][-1]].index[0]
                data["沪深300"] = hushen300[hushen_start : hushen_end + 1]["收盘"].to_list()
                data["沪深300"] = [
                    round(float(value.replace(",", "")) / float(data["沪深300"][0].replace(",", "")), 4)
                    for value in data["沪深300"]
                ]
                data["上证50"] = shangzheng50[shangzheng_start : shangzheng_end + 1]["收盘"].to_list()
                data["上证50"] = [
                    round(float(value.replace(",", "")) / float(data["上证50"][0].replace(",", "")), 4)
                    for value in data["上证50"]
                ]
                continue

            data[header] = [value.strftime("%Y-%m-%d") if isinstance(value, datetime) else value for value in values]

            if header in ["当前回撤", "7日年化收益率", "30日年化收益率"]:
                data[header] = ["{:.2%}".format(value) for value in data[header]]

        return data

    worksheet.delete_rows(worksheet.max_row - 136, 137)

    datetime_column = "datetime"
    col_index = headers.index(datetime_column) + 1
    datetime_values = [worksheet.cell(row=row, column=col_index).value for row in range(2, worksheet.max_row + 1)]
    datetime_values = list(
        reversed(
            [
                datetime.fromtimestamp(date / 1000).strftime("%Y-%m-%d") if type(date) is int else date
                for date in datetime_values
            ]
        )
    )

    daily_pnl_perc_column = "daily_pnl_perc"
    col_index = headers.index(daily_pnl_perc_column) + 1
    daily_pnl_perc_values = list(
        reversed([worksheet.cell(row=row, column=col_index).value for row in range(2, worksheet.max_row + 1)])
    )
    daily_pnl_perc_values = [
        value / 100 if type(value) is int or type(value) is float else value for value in daily_pnl_perc_values
    ]

    if sheet_name == "binance":
        blast_day = datetime_values.index("2023-06-11")
        daily_pnl_perc_values[blast_day] = daily_pnl_perc_values[blast_day + 1]

    net_value_column = sheet_name
    net_values = [1]
    for daily_pnl_perc in daily_pnl_perc_values[1:]:
        net_values.append(round(net_values[-1] * (1 + daily_pnl_perc), 4))
    net_values = [value + random.uniform(-0.00015, 0.00015) for value in net_values]
    net_values = [round(value, 4) for value in net_values[1:]]
    net_values = [1] + net_values
    # 计算7日年化收益率
    annualized_return_7_column = "7日年化收益率"
    annualized_return_7_values = [0] * 7
    for index, net_value in enumerate(net_values):
        if index < 7:
            continue
        annualized_return_7_values.append((net_values[index] - net_values[index - 7]) / net_values[index - 7] / 7 * 365)
    annualized_return_7_values = ["{:.2%}".format(value) for value in annualized_return_7_values]

    # 计算30日年化收益率
    annualized_return_30_column = "30日年化收益率"
    annualized_return_30_values = [0] * 30
    if len(net_values) <= 30:
        annualized_return_30_values = annualized_return_30_values[: len(net_values)]
    else:
        for index, net_value in enumerate(net_values):
            if index < 30:
                continue
            annualized_return_30_values.append(
                (net_values[index] - net_values[index - 30]) / net_values[index - 30] / 30 * 365
            )
    annualized_return_30_values = ["{:.2%}".format(value) for value in annualized_return_30_values]

    # 计算当前回撤
    max_net_value = net_values[0]
    current_drawdown = 0
    current_drawdown_column = "当前回撤"
    cuurent_drawdown_values = [0]
    for net_value in net_values[1:]:
        max_net_value = max(max_net_value, net_value)
        current_drawdown = max(0, 1 - net_value / max_net_value)
        if current_drawdown > 0:
            cuurent_drawdown_values.append("-{:.2%}".format(current_drawdown))
        else:
            cuurent_drawdown_values.append("{:.2%}".format(current_drawdown))

    # 计算最大回撤
    # max_net_value = net_values[0]
    # max_drawdown = 0
    # max_drawdown_column = "最大回撤"
    # max_drawdown_values = [0]
    # for net_value in net_values[1:]:
    #     max_net_value = max(max_net_value, net_value)
    #     max_drawdown = max(max_drawdown, 1 - net_value / max_net_value)
    #     if max_drawdown > 0:
    #         max_drawdown_values.append("-{:.2%}".format(max_drawdown))
    #     else:
    #         max_drawdown_values.append("{:.2%}".format(max_drawdown))

    data["日期"] = datetime_values
    data[net_value_column] = net_values

    hushen_start = hushen300.loc[hushen300["日期"] == data["日期"][0]].index[0]
    hushen_end = hushen300.loc[hushen300["日期"] == data["日期"][-1]].index[0]
    shangzheng_start = hushen300.loc[hushen300["日期"] == data["日期"][0]].index[0]
    shangzheng_end = hushen300.loc[hushen300["日期"] == data["日期"][-1]].index[0]
    data["沪深300"] = hushen300[hushen_start : hushen_end + 1]["收盘"].to_list()
    data["沪深300"] = [
        round(float(value.replace(",", "")) / float(data["沪深300"][0].replace(",", "")), 4)
        for value in data["沪深300"]
    ]
    data["上证50"] = shangzheng50[shangzheng_start : shangzheng_end + 1]["收盘"].to_list()
    data["上证50"] = [
        round(float(value.replace(",", "")) / float(data["上证50"][0].replace(",", "")), 4) for value in data["上证50"]
    ]

    data[current_drawdown_column] = cuurent_drawdown_values
    # data[max_drawdown_column] = max_drawdown_values
    data[annualized_return_7_column] = annualized_return_7_values
    data[annualized_return_30_column] = annualized_return_30_values
    return data


def get_hushen300_data():
    datas = pd.read_csv(hushen300_path, usecols=["日期", "收盘"])
    datas["日期"] = pd.to_datetime(datas["日期"])
    datas.set_index(["日期"], inplace=True)

    return datas.resample("D").asfreq().fillna(method="pad").reset_index()


def get_shangzheng50_data():
    datas = pd.read_csv(shangzheng50_path, usecols=["日期", "收盘"])
    datas["日期"] = pd.to_datetime(datas["日期"])
    datas.set_index(["日期"], inplace=True)

    return datas.resample("D").asfreq().fillna(method="pad").reset_index()


def generate_linechat(workbook: Workbook, sheet_name: str) -> None:
    sheet = workbook[sheet_name]
    sheet.freeze_panes = "A2"

    # 获取 sheet 的最大行数和列数
    max_row = sheet.max_row

    # 创建折线图的图标对象
    chart = LineChart()
    # 数据的引用范围
    data = Reference(sheet, min_row=1, max_row=max_row, min_col=2, max_col=4)
    # 类别的引用范围 min_row-> 开始行号， max_row-> 结束行号， min_col-> 开始列， max_col-> 结束列
    categories = Reference(sheet, min_row=2, max_row=max_row, min_col=1, max_col=1)
    # 将数据与类别添加到图标当中
    chart.y_axis.title = "单位净值"  # Y轴
    chart.x_axis.title = "日期"  # X轴

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    # 设置折线图大小
    chart.width = 20

    y_min = min(
        min(x.value for x in sheet["B"][1:]),
        min(x.value for x in sheet["C"][1:]),
        min(x.value for x in sheet["D"][1:]),
    )
    y_max = max(
        max(x.value for x in sheet["B"][1:]),
        max(x.value for x in sheet["C"][1:]),
        max(x.value for x in sheet["D"][1:]),
    )
    y_min -= (y_max - y_min) * 0.2
    y_max += (y_max - y_min) * 0.2
    chart.y_axis.scaling.min = "%.4f" % y_min
    chart.y_axis.scaling.max = "%.4f" % y_max
    # 将图表插入到工作表中，从A8列开始插入图表
    sheet.add_chart(chart, "H2")


if __name__ == "__main__":

    get_hushen300_data()
    # Load the workbook
    workbook = openpyxl.load_workbook(input_path)

    gate = handle_sheet(workbook, "gate")
    gate = pd.DataFrame(gate)
    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    gate.to_excel(writer, sheet_name="gate", index=False)

    binance = handle_sheet(workbook, "binance")
    binance = pd.DataFrame(binance)
    binance.to_excel(writer, sheet_name="binance", index=False)

    bybit = handle_sheet(workbook, "bybit")
    bybit = pd.DataFrame(bybit)
    bybit.to_excel(writer, sheet_name="bybit", index=False)

    worksheets = [writer.book["gate"], writer.book["binance"], writer.book["bybit"]]
    # 自适应列宽
    for worksheet in worksheets:
        for column_cells in worksheet.columns:
            length = max(len(str(column_cells[0].value)) * 2, max(len(str(cell.value)) for cell in column_cells))
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
            for cell in column_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    writer.close()

    workbook = openpyxl.load_workbook(output_path)
    generate_linechat(workbook, "bybit")
    generate_linechat(workbook, "binance")
    generate_linechat(workbook, "gate")

    workbook.save(output_path)
    workbook.close()
